#!/usr/bin/env python3
"""
Operational Information Parsing Engine
=======================================

Reads a trip-support / ops-preferences document (PDF -- including
scanned or "flattened" image-only PDFs, which is what ForeFlight's
exported/completed forms typically are) and classifies its content into
a FIXED set of 19 "Operational Information" subjects, then writes an
.xlsx in the Subject / Comments grid format used by the ops team.

Two of those subjects carry fixed boilerplate that is NOT derived from
the source PDF -- see add_static_content() and the constants above it:
  - "High Profile Notes": a fixed new-customer onboarding note, inserted
    verbatim (with XXX/XXXXX placeholders) on every run.
  - "General Info": gets a fixed "Jeppesen FF Account Exec: (xx Add Name
    xx)" line appended, plus the Aircraft Details "what types of flight
    do you operate" answer is rendered as the full fixed checklist with
    the customer's selection(s) marked, rather than just the raw value.

Why vision instead of OCR/text extraction:
    The reference document ("Trip Support Service Preferences") is a
    flattened PDF -- every page is a single embedded image with zero
    extractable text. Answers are conveyed largely through checkbox
    fill-state (checked vs. unchecked), which plain OCR resolves
    unreliably. Each page is instead sent to Claude as an image and
    read the way a person would.

Pipeline
--------
    1. render_pages()        PDF -> one PNG per page (PyMuPDF)
    2. extract_page()        page image -> structured Q&A/table JSON
                              (Claude vision, one call per page)
    3. map_to_subjects()     structured Q&A -> {subject: [lines]}
                              via SECTION_MAP (see below)
    4. write_workbook()      {subject: comments} -> styled .xlsx
                              matching the Operational Information grid

Nothing is auto-extracted by default
-------------------------------------
FIELD_WHITELIST is the single source of truth for which fields the
engine is allowed to pull out of the source document. A field -- even
if clearly answered on the page -- is only extracted if its exact label
appears in FIELD_WHITELIST for that section. This is enforced twice:
once in the prompt sent to Claude (it's told the exact approved field
list and told to ignore everything else), and again in code via
is_whitelisted(), which hard-drops any item the model returns that
isn't on the list. Dropped fields are reported in the log, not silently
discarded. To let a new field flow through, add it to FIELD_WHITELIST.

Extending to new source-document types
---------------------------------------
SECTION_MAP and FIELD_WHITELIST both need entries. SECTION_MAP maps a
source document's own section headings onto the fixed SUBJECTS list
below; FIELD_WHITELIST says which fields under that heading are
approved to extract. Sections with no SECTION_MAP entry, and fields
with no FIELD_WHITELIST entry, are dropped (reported in the log), and
subjects with no matching source content are left blank in the output
-- exactly like the blank template in the target grid.

Requirements
------------
    pip install pymupdf anthropic openpyxl --break-system-packages
    export ANTHROPIC_API_KEY=...

Usage
-----
    python parsing_engine.py "input.pdf" -o "output.xlsx"

    # Re-run formatting/mapping without re-paying for vision calls,
    # using the cached raw extraction from a previous run:
    python parsing_engine.py "input.pdf" -o "output.xlsx" --use-cache
"""

import argparse
import base64
import json
import re
import sys
from pathlib import Path

# ---------------------------------------------------------------------------
# 1. Fixed output subjects -- must match the Operational Information grid
# ---------------------------------------------------------------------------
SUBJECTS = [
    "Company Details",
    "Communications",
    "High Profile Notes",
    "General Info",
    "Flight Following",
    "Ground Handling",
    "Permits",
    "Hotels",
    "Fuel",
    "Customs Information",
    "Catering",
    "Ground Transport",
    "Sponsor Information",
    "Emergency Procedures",
    "FP - General Procedures",
    "FP - Options/Formats",
    "FP - Aircraft Databases",
    "FP - Alternate Airports",
    "FP - Preferred Routing",
    "FP - Weather Packages",
    "FP - Remarks",
]

# ---------------------------------------------------------------------------
# 2. Section -> Subject mapping.
#    Key = a substring to match (lower-cased) against the source
#    document's own section heading. First match wins.
#    This table is tuned to the "Jeppesen ForeFlight Trip Support
#    Service Preferences" form (Ver 2.2). Add rows here for other
#    source-document types -- nothing else in the engine needs to change.
# ---------------------------------------------------------------------------
SECTION_MAP = {
    "company details": "Company Details",
    "aircraft details": "General Info",  # only the flight-type field is whitelisted
    "flight monitoring": "Flight Following",
    "ground handling": "Ground Handling",
    "permits": "Permits",
    "hotels": "Hotels",
    "fuel": "Fuel",
    "customs": "Customs Information",
    "catering": "Catering",
    "ground transport": "Ground Transport",
    "general comments": "FP - Remarks",
}

# ---------------------------------------------------------------------------
# 2b. Field whitelist -- the ONLY fields the engine is allowed to pull out
#     of each section. Nothing is auto-extracted unless its exact question/
#     field label is listed here. Anything else on the page -- answered or
#     not -- is ignored and reported in the log as "not on whitelist".
#
#     To let a new field flow through: add its label here (matched
#     case-insensitively, so wording doesn't have to be pixel-perfect).
#     This is the single source of truth: both the vision prompt sent to
#     Claude and the hard post-extraction filter are built from this dict,
#     so a field can't slip through by the model volunteering it.
# ---------------------------------------------------------------------------
FIELD_WHITELIST = {
    # Feeds BOTH the "Company Details" and "Communications" subjects --
    # see MULTI_SUBJECT_FORMATTERS, format_company_details_section(), and
    # format_communications_section(). Core Service Preferences and
    # Customer Tools remain NOT listed here -- their fields are still not
    # approved to auto-extract anywhere.
    "company details": [
        "Company Name", "Company Address", "Home Base(s) ICAO(s)",
        "Operations team?",  # Q3
        "Ops team email",    # Q4
        "Key contact",       # Q5 -- one item per contact row
        "Billing contact",   # Q6 -- one item per contact row
    ],
    # Aircraft Details is trimmed to only the flight-type question for
    # the same reason (feeds General Info, not Company Details).
    "aircraft details": [
        "Flight types operated",
    ],
    # Only these two raw fields feed the derived Flight Following /
    # Flight Monitoring lines -- see format_flight_monitoring_section().
    "flight monitoring": [
        "Movement messages opt-out checkbox",  # Q1
        "Send movements as an added service",  # Q2b
    ],
    "permits": [
        "Countries to avoid overflying", "Block (seasonal) permissions held",
    ],
    # Fixed 10-field list, rendered in this exact order by
    # format_customs_section() -- order here IS the output order.
    "customs": [
        "USA Customs Notification", "eAPIS (USA)", "eAPIS (other countries)",
        "UK Air Passenger Duty (APD)", "UK General Avtn Report (GAR)",
        "USA TSA Waiver held", "USA CBP Border Overflight Exemption held",
        "Visa Waiver Program Signatory Carrier", "Third Country Operator (TCO)",
        "APIS Carrier code",
    ],
    # Only these two feed the derived towbar lines -- see
    # format_ground_handling_section(). Split into two fields (rather than
    # one combined "Yes -- N680BA" field) so the formatter can tell "Yes,
    # no registration given" apart from "No".
    "ground handling": [
        "Towbar carried checkbox",       # Q2 Yes/No
        "Aircraft tails with towbars",   # the listed registration(s), if any
    ],
    # Only this field feeds the derived Crew/Passenger Catering lines --
    # see format_catering_section().
    "catering": [
        "Catering required for",
    ],
    # Only this field feeds the Fuel section's single output line -- see
    # format_fuel_section().
    "fuel": [
        "Preferred fuel provider",
    ],
    # Q4 (Hotel Membership Reward Cards) intentionally excluded -- not
    # part of the Hotels section's defined output. Q11 is split into two
    # fields (Yes/No + free-text criteria) so they can be rendered as two
    # independent lines -- see format_hotels_section().
    "hotels": [
        "Hotels required for", "Hotel options sent prior to booking",
        "Preferred hotel chain", "Hotel price range", "Room type preference",
        "Minimum hotel star rating", "Location for stays of 1 night",
        "Location for stays of more than 1 night", "Payment preference",
        "Book room the night before for early AM arrival",
        "Early AM arrival criteria",
    ],
    "ground transport": [
        "Transport required for", "Crew pickup time after aircraft arrival",
        "Crew drop-off time before departure", "Mode of transport for crew",
        "Mode of transport for passengers", "Preferred car rental companies",
        "Preferred car type",
    ],
    "general comments": [
        "Challenges with other Trip Support Providers",
    ],
}


# Subjects this source-document type is not expected to cover at all
# (e.g. a company preferences form has no per-trip flight-plan detail).
# Left blank in output rather than guessed at.
NOT_COVERED_BY_THIS_DOC_TYPE = {
    "Sponsor Information",
    "Emergency Procedures",
    "FP - General Procedures",
    "FP - Options/Formats",
    "FP - Aircraft Databases",
    "FP - Alternate Airports",
    "FP - Preferred Routing",
    "FP - Weather Packages",
}

# ---------------------------------------------------------------------------
# 3. Static / boilerplate content -- NOT derived from the source PDF.
#    Injected into every run's output by add_static_content() below.
#    Edit the text here (not in map_to_subjects) to change what gets added.
# ---------------------------------------------------------------------------

# Always inserted into "High Profile Notes", verbatim, on every run.
# Placeholders (XXX / XXXXX / 2026) are intentional -- fill in per customer
# after the sheet is generated.
HIGH_PROFILE_NOTES_TEMPLATE = """\
** NEW CUSTOMER - XXX 2026 *****ONBOARDING TEAM ACCOUNT (Jodi, Michal, Scott, Ruby, Bipin, Robert)

Credit Card Customer - Advance deposit required.
(If CC customer, click “Visa” under Payment Methods and leave card # blank). This will trigger the planners to get a CC deposit.

FFD acct under 'XXXXX' (linked)"""

# Always appended to "General Info" on every run. The Account Exec name
# comes from the Salesperson selected in the intake UI (passed through
# map_to_subjects()/add_static_content() as `salesperson`); if none is
# given -- e.g. running this script standalone from the CLI -- it falls
# back to a manual fill-in placeholder.
GENERAL_INFO_ACCOUNT_EXEC_PLACEHOLDER = ">> Add Name <<"


def _general_info_static_lines(salesperson: str | None = None) -> list[str]:
    name = (salesperson or "").strip() or GENERAL_INFO_ACCOUNT_EXEC_PLACEHOLDER
    return [f"Jeppesen FF Account Exec: {name}"]

# The fixed option set for Aircraft Details Q2 ("What types of flight do
# you operate?"). Always rendered in full with the customer's selection(s)
# marked, rather than only listing the selected value(s).
FLIGHT_TYPE_OPTIONS = [
    "Private Non-Commercial",
    "Non-Scheduled Commercial (Charter)",
    "Scheduled Commercial",
    "Diplomatic (Govt/Military)",
    "Cargo",
    "Air Ambulance",
    "Ferry Delivery Flights",
    "Other (List)",
]

# Label(s) the extraction step may use for this question -- matched
# case-insensitively as a substring against each extracted item's label.
FLIGHT_TYPE_ITEM_LABELS = ("flight type", "flight(s) do you operate", "types of flight")

# Some fields need extraction to report a state EVEN WHEN the box is
# unchecked (the default "skip unanswered/unchecked" rule doesn't apply).
# heading -> {field label substring: extra instruction text}
SPECIAL_FIELD_INSTRUCTIONS = {
    "Movement messages opt-out checkbox": (
        'This is a single opt-out checkbox ("If you do not want us to provide this '
        'service, please check here"), not a Yes/No pair -- it is normal for it to '
        'be unchecked. Unlike other checkbox fields, ALWAYS report its state: '
        'value must be exactly "Checked" if the box is filled/checked, or '
        '"Unchecked" if it is empty. Do not omit this field even when unchecked.'
    ),
    "Key contact": (
        'The table has separate "Tel" (office/home phone) and "Mobile" columns. '
        'Format each row\'s value as "<Name>, <Role>, <Email>[, Office: <Tel>]'
        '[, Mobile: <Mobile>]" -- label each number with the column it came from, '
        'in that order, and omit either label entirely if that column is blank on '
        'that row. Emit one item per filled-in contact row.'
    ),
    "Billing contact": (
        'Same table format as "Key contact": separate "Tel" (office/home phone) '
        'and "Mobile" columns. Format each row\'s value as "<Name>, <Role>, <Email>'
        '[, Office: <Tel>][, Mobile: <Mobile>]", omitting either label if that '
        'column is blank. Emit one item per filled-in contact row.'
    ),
}

MODEL = "claude-sonnet-5"


def _build_extraction_prompt() -> str:
    """Built from FIELD_WHITELIST so the model is only ever asked for
    fields that are explicitly approved to auto-extract -- it is told,
    per section, exactly which fields it may report and instructed to
    ignore everything else on the page, answered or not."""
    checklist = []
    for heading, fields in FIELD_WHITELIST.items():
        field_list = "; ".join(fields)
        checklist.append(f'  - "{heading.title()}": {field_list}')
    checklist_text = "\n".join(checklist)

    special_notes = "\n".join(
        f'  - "{label}": {instr}' for label, instr in SPECIAL_FIELD_INSTRUCTIONS.items()
    )

    return f"""You are reading one page of a filled-out aviation trip-support \
preferences form (a scanned/flattened PDF page, provided as an image).

You may ONLY extract the specific fields listed below, grouped by the section \
heading they fall under. Do NOT extract any other field, question, or checkbox on \
the page, even if it is clearly answered -- if it is not in this list, leave it out:

{checklist_text}

For each approved field that IS answered on this page:
  - Use the exact field name from the list above as "label".
  - For a checkbox field, report only the option(s) that are checked (filled box /
    checkmark) as the "value" -- do not list unchecked options.
  - For a table with repeating rows (e.g. contacts, accounts), emit one item per row.
If an approved field is blank/unanswered on this page, omit it entirely.
Do not invent, rename, or reword field labels -- match them to the list above.

Special handling notes for specific fields (override the general rule above):
{special_notes}

Return ONLY valid JSON (no markdown fences, no commentary) matching this shape:

{{
  "sections": [
    {{
      "heading": "<one of the section headings listed above>",
      "items": [
        {{"label": "<one of the approved field names for that section>", "value": "<answered value>"}}
      ]
    }}
  ]
}}

If nothing approved is answered on this page, return {{"sections": []}}.
"""


EXTRACTION_PROMPT = _build_extraction_prompt()


def render_pages(pdf_path: Path, dpi: int = 200) -> list[bytes]:
    """PDF -> list of PNG bytes, one per page."""
    import fitz  # PyMuPDF

    images = []
    with fitz.open(pdf_path) as doc:
        for page in doc:
            pix = page.get_pixmap(dpi=dpi)
            images.append(pix.tobytes("png"))
    return images


def extract_page(png_bytes: bytes, client, model: str = MODEL) -> dict:
    """Send one page image to Claude vision, get back structured JSON."""
    b64 = base64.standard_b64encode(png_bytes).decode("ascii")
    resp = client.messages.create(
        model=model,
        max_tokens=2000,
        messages=[
            {
                "role": "user",
                "content": [
                    {
                        "type": "image",
                        "source": {
                            "type": "base64",
                            "media_type": "image/png",
                            "data": b64,
                        },
                    },
                    {"type": "text", "text": EXTRACTION_PROMPT},
                ],
            }
        ],
    )
    text = "".join(block.text for block in resp.content if block.type == "text")
    text = re.sub(r"^```(?:json)?|```$", "", text.strip(), flags=re.MULTILINE).strip()
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        print(f"  [warn] could not parse model output as JSON, skipping page:\n{text[:300]}", file=sys.stderr)
        return {"sections": []}


def extract_document(pdf_path: Path, cache_path: Path, use_cache: bool = False) -> list[dict]:
    """Run render_pages + extract_page over the whole document, with a
    JSON cache so re-runs of the mapping/output step don't re-call the
    API."""
    if use_cache and cache_path.exists():
        print(f"Using cached extraction: {cache_path}")
        return json.loads(cache_path.read_text())["sections"]

    import anthropic

    client = anthropic.Anthropic()
    all_sections = []
    pages = render_pages(pdf_path)
    for i, png in enumerate(pages, start=1):
        print(f"Extracting page {i}/{len(pages)}...")
        result = extract_page(png, client)
        for section in result.get("sections", []):
            section["page"] = i
            all_sections.append(section)

    cache_path.write_text(json.dumps({"sections": all_sections}, indent=2))
    return all_sections


def resolve_subject(heading: str) -> str | None:
    key = heading.strip().lower()
    for pattern, subject in SECTION_MAP.items():
        if pattern in key:
            return subject
    return None


def _normalize(s: str) -> str:
    return re.sub(r"[^a-z0-9]+", " ", s.lower()).strip()


def is_whitelisted(heading: str, label: str) -> bool:
    """Hard backstop, independent of the prompt: True only if `label` is
    an explicitly approved field for `heading` in FIELD_WHITELIST. Applied
    to every extracted item regardless of what the model returned, so a
    field can never reach the output without being named here first."""
    heading_key = heading.strip().lower()
    allowed_fields = None
    for pattern, fields in FIELD_WHITELIST.items():
        if pattern in heading_key:
            allowed_fields = fields
            break
    if allowed_fields is None:
        return False

    norm_label = _normalize(label)
    for allowed in allowed_fields:
        norm_allowed = _normalize(allowed)
        if norm_label == norm_allowed or norm_allowed in norm_label or norm_label in norm_allowed:
            return True
    return False


def format_flight_types(selected_value: str) -> str:
    """Render 'What types of flight do you operate?' as a single line
    listing only the option(s) the customer actually selected, matched
    against the fixed option set so wording stays consistent with the
    source form. selected_value may contain comma-separated selections."""
    selected = [s.strip().lower() for s in selected_value.split(",") if s.strip()]

    def is_checked(option: str) -> bool:
        base = option.split(" (")[0].strip().lower()  # e.g. "Other" from "Other (List)"
        return any(sel == base or sel.startswith(base) or base.startswith(sel) for sel in selected)

    matched = [option for option in FLIGHT_TYPE_OPTIONS if is_checked(option)]
    values = matched if matched else [v.strip() for v in selected_value.split(",") if v.strip()]
    return "Customer operates: " + ", ".join(values)


_EMPTY_VALUES = {"", "none", "none listed", "none held", "n/a", "no"}


def _strip_yes_prefix(text: str) -> str:
    """'Yes — Iran, Russia, ...' -> 'Iran, Russia, ...'. Source checkbox
    answers are often recorded as an affirmative plus the actual list; we
    only want the list itself in the output."""
    m = re.match(r"^\s*yes\b[\s\-–—:,]*", text, flags=re.IGNORECASE)
    return text[m.end():].strip() if m else text.strip()


def format_permits_section(kept_items: list[tuple[str, str]]) -> list[str]:
    """Permits always opens with a fixed statement, then two lines that
    fall back to a default sentence unless the customer actually entered
    something in that field -- in which case show their entry (with any
    leading "Yes" affirmation stripped) instead of the default."""
    values: dict[str, str] = {}
    for label, value in kept_items:
        norm_label = _normalize(label)
        for canonical in FIELD_WHITELIST["permits"]:
            norm_canonical = _normalize(canonical)
            if norm_label == norm_canonical or norm_canonical in norm_label or norm_label in norm_canonical:
                values[canonical] = value
                break

    lines = ["JeppFF will arrange all Permits (unless otherwise indicated)."]

    block_value = values.get("Block (seasonal) permissions held", "").strip()
    if block_value.lower() in _EMPTY_VALUES:
        lines.append("They do not hold any block (seasonal) permits.")
    else:
        lines.append(_strip_yes_prefix(block_value))

    avoid_value = values.get("Countries to avoid overflying", "").strip()
    if avoid_value.lower() in _EMPTY_VALUES:
        lines.append("No countries specified to avoid when planning flights.")
    else:
        lines.append(f"Countries to Avoid: {_strip_yes_prefix(avoid_value)}")

    return lines


def format_catering_section(kept_items: list[tuple[str, str]]) -> list[str]:
    """Two derived lines from 'Catering required for' (which lists Crew,
    Passengers, both, or Neither): each person-group gets its own
    Jeppesen-arranges / Not Required line."""
    raw_value = ""
    for label, value in kept_items:
        if "catering required for" in _normalize(label) or "required for" in _normalize(label):
            raw_value = value
            break
    selected = [s.strip().lower() for s in raw_value.split(",") if s.strip()]
    crew_selected = any(s.startswith("crew") for s in selected)
    pax_selected = any(s.startswith("passenger") or s.startswith("pax") for s in selected)

    return [
        "Crew Catering: " + ("Jeppesen FF to arrange" if crew_selected else "Not Required"),
        "Passenger Catering: " + ("Jeppesen FF to arrange" if pax_selected else "Not Required"),
    ]


def format_company_details_section(kept_items: list[tuple[str, str]]) -> list[str]:
    """Company Name and Company Address pass straight through. The home
    base field only gets shown if it's a single, valid 4-character
    alphanumeric ICAO code -- e.g. "KADS, EGLL" or a stray note wouldn't
    pass and gets left out rather than shown malformed."""
    values: dict[str, str] = {}
    for label, value in kept_items:
        norm = _normalize(label)
        if "company name" in norm:
            values["name"] = value
        elif "company address" in norm:
            values["address"] = value
        elif "home base" in norm or "icao" in norm:
            values["home_base"] = value

    lines = []
    if values.get("name", "").strip():
        lines.append(f"Company Name: {values['name'].strip()}")
    if values.get("address", "").strip():
        lines.append(f"Company Address: {values['address'].strip()}")
    home_base = values.get("home_base", "").strip()
    if re.fullmatch(r"[A-Za-z0-9]{4}", home_base):
        lines.append(f"Home Base: {home_base.upper()}")
    return lines


def format_communications_section(kept_items: list[tuple[str, str]]) -> list[str]:
    """Draws from the SAME Company Details fields as
    format_company_details_section(), just picking out different ones.
    Priority order (each only contributes if the higher priority didn't):
      1. Q3 (Operations team?) = Yes -> show Q4's group/individual email,
         or nothing at all if Q4 was left blank.
      2. All Q5 key contacts (Main Contact/Managers/Chief Pilot), one line each.
      3. All Q6 billing contacts, one line each.
    All three priorities can independently contribute -- "priority" here
    means the fallback order Q3/Q4 imply, not an either/or across all three."""
    ops_team_value = ""
    ops_email_value = ""
    key_contacts: list[str] = []
    billing_contacts: list[str] = []
    for label, value in kept_items:
        norm = _normalize(label)
        if "operations team" in norm:
            ops_team_value = value
        elif "ops team email" in norm or "main ops email" in norm:
            ops_email_value = value
        elif "key contact" in norm:
            key_contacts.append(value)
        elif "billing contact" in norm:
            billing_contacts.append(value)

    entries = []

    # Priority 1
    if ops_team_value.strip().lower().startswith("yes") and ops_email_value.strip():
        entries.append(f"Ops Team Contact: {ops_email_value.strip()}")

    # Priority 2
    for contact in key_contacts:
        entries.append(f"Key Contact: {contact}")

    # Priority 3
    for contact in billing_contacts:
        entries.append(f"Billing Contact: {contact}")

    # Blank line between each contact -- "" becomes an empty line once the
    # caller joins this list with "\n".
    lines = []
    for i, entry in enumerate(entries):
        if i > 0:
            lines.append("")
        lines.append(entry)
    return lines


# Some headings need to feed TWO different output subjects from the same
# whitelisted fields (rather than the usual one heading -> one subject).
# heading pattern -> [(target subject, formatter), ...]. When a heading
# matches here, it bypasses SECTION_MAP/SPECIAL_SECTION_FORMATTERS
# entirely -- each formatter gets the full whitelisted item list and
# picks out whatever it needs.
MULTI_SUBJECT_FORMATTERS = {
    "company details": [
        ("Company Details", format_company_details_section),
        ("Communications", format_communications_section),
    ],
}


def format_ground_handling_section(kept_items: list[tuple[str, str]]) -> list[str]:
    """From Q2 ('Do you carry a Towbar with each aircraft?') and the
    registrations text field:
      Yes + registration(s) given -> one "<tail> - Towbar Carried" line each
      Yes + no registration given -> a flag line asking for the tail
      No                          -> no lines at all
    """
    checkbox_value = ""
    tails_value = ""
    for label, value in kept_items:
        norm = _normalize(label)
        if "towbar carried" in norm or "carries towbar" in norm:
            checkbox_value = value
        elif "aircraft tails" in norm or "tails with towbars" in norm:
            tails_value = value

    if not checkbox_value.strip().lower().startswith("yes"):
        return []

    tails = [t.strip() for t in tails_value.split(",") if t.strip()]
    if tails:
        return [f"{tail} - Towbar Carried" for tail in tails]
    return [
        "Operator has stated towbar carried but did not provide aircraft "
        "registration (REMOVE FOR SINGLE TAIL OPERATOR)"
    ]


def format_hotels_section(kept_items: list[tuple[str, str]]) -> list[str]:
    """Fixed-order Hotels block. Crew/Passenger Hotel are derived from
    the 'Hotels required for' checkbox; Quote Required is derived from
    the 'send options prior to booking' Yes/No. Everything else passes
    through as the customer's raw answer, defaulting to "N/A" when
    unanswered -- except the early-arrival criteria line, which is left
    truly blank (not "N/A") when the customer didn't provide one."""
    values: dict[str, object] = {}
    for label, value in kept_items:
        norm_label = _normalize(label)
        for canonical in FIELD_WHITELIST["hotels"]:
            norm_canonical = _normalize(canonical)
            if norm_label == norm_canonical or norm_canonical in norm_label or norm_label in norm_canonical:
                if canonical == "Preferred hotel chain":
                    values.setdefault(canonical, []).append(value)
                else:
                    values[canonical] = value
                break

    def raw(field: str, default: str = "N/A") -> str:
        v = values.get(field)
        if not v or not str(v).strip():
            return default
        return str(v)

    required_for = str(values.get("Hotels required for", ""))
    selected = [s.strip().lower() for s in required_for.split(",") if s.strip()]
    crew_selected = any(s.startswith("crew") for s in selected)
    pax_selected = any(s.startswith("passenger") or s.startswith("pax") for s in selected)

    quote_value = str(values.get("Hotel options sent prior to booking", "")).strip().lower()
    if quote_value.startswith("yes"):
        quote_line = "Jeppesen FF to send 3 quotes before booking"
    elif quote_value.startswith("no"):
        quote_line = "Not Required"
    else:
        quote_line = "N/A"

    chains = values.get("Preferred hotel chain", [])

    return [
        "Crew Hotel: " + ("Jeppesen FF to arrange" if crew_selected else "Customer to do own"),
        "Passenger Hotel: " + ("Jeppesen FF to arrange" if pax_selected else "Customer to do own"),
        f"Quote Required: {quote_line}",
        "Preferred Chain(s): " + (", ".join(chains) if chains else "N/A"),
        f"Budget: {raw('Hotel price range')}",
        f"Room Type: {raw('Room type preference')}",
        f"Minimum Hotel Star Rating: {raw('Minimum hotel star rating')}",
        f"Location for stay of 1 night or less: {raw('Location for stays of 1 night')}",
        f"Location for stay more than 1 day: {raw('Location for stays of more than 1 night')}",
        f"Payment preference: {raw('Payment preference')}",
        f"For early AM arrival, book rooms for night before: {raw('Book room the night before for early AM arrival')}",
        f"Provide criteria: {raw('Early AM arrival criteria', default='')}",
    ]


def format_ground_transport_section(kept_items: list[tuple[str, str]]) -> list[str]:
    """Fixed-order Ground Transport block. Crew/Passenger are derived
    from the 'Transport required for' checkbox; everything else passes
    through as the customer's raw answer, defaulting to "N/A" when
    unanswered."""
    values: dict[str, str] = {}
    for label, value in kept_items:
        norm_label = _normalize(label)
        for canonical in FIELD_WHITELIST["ground transport"]:
            norm_canonical = _normalize(canonical)
            if norm_label == norm_canonical or norm_canonical in norm_label or norm_label in norm_canonical:
                values[canonical] = value
                break

    def raw(field: str, default: str = "N/A") -> str:
        v = values.get(field, "")
        return v if v.strip() else default

    required_for = values.get("Transport required for", "")
    selected = [s.strip().lower() for s in required_for.split(",") if s.strip()]
    crew_selected = any(s.startswith("crew") for s in selected)
    pax_selected = any(s.startswith("passenger") or s.startswith("pax") for s in selected)

    return [
        "Crew: " + ("Jeppesen FF to arrange" if crew_selected else "customer to arrange own"),
        "Passenger: " + ("Jeppesen FF to arrange" if pax_selected else "customer to arrange own"),
        f"Pick crew up after ETA: {raw('Crew pickup time after aircraft arrival')}",
        f"Drop crew off before ETD: {raw('Crew drop-off time before departure')}",
        f"Mode for crew: {raw('Mode of transport for crew')}",
        f"Mode for passengers: {raw('Mode of transport for passengers')}",
        f"Preferred Rental Car Company: {raw('Preferred car rental companies')}",
        f"Preferred Car type: {raw('Preferred car type')}",
    ]


def format_fuel_section(kept_items: list[tuple[str, str]]) -> list[str]:
    """Fuel is trimmed to a single line: the customer's preferred fuel
    provider, defaulting to N/A when left unanswered."""
    value = ""
    for label, v in kept_items:
        if "preferred fuel provider" in _normalize(label):
            value = v
            break
    return [f"Preferred Fuel Provider: {value.strip() if value.strip() else 'N/A'}"]


def format_aircraft_details_section(kept_items: list[tuple[str, str]]) -> list[str]:
    """Aircraft Details is trimmed to just the flight-type question --
    render it as a single clean line with no [Aircraft Details] heading
    wrapper, since General Info should show nothing else from this
    section."""
    lines = []
    for label, value in kept_items:
        if any(marker in label.lower() for marker in FLIGHT_TYPE_ITEM_LABELS):
            lines.append(format_flight_types(value))
    return lines


# These 5 customs fields are rendered as "who arranges it" rather than
# raw Yes/No: Yes -> Jeppesen handles it, No -> the customer handles it
# themselves. The other 5 customs fields keep the customer's raw answer.
CUSTOMS_ARRANGE_FIELDS = {
    "USA Customs Notification",
    "eAPIS (USA)",
    "eAPIS (other countries)",
    "UK Air Passenger Duty (APD)",
    "UK General Avtn Report (GAR)",
}


def format_customs_section(kept_items: list[tuple[str, str]]) -> list[str]:
    """Render the customs fields in the fixed order given in
    FIELD_WHITELIST['customs'] (not extraction order), one per line, with
    no [Customs] heading wrapper. Any field not answered on the page
    defaults to "N/A" rather than being omitted. For CUSTOMS_ARRANGE_FIELDS,
    the raw Yes/No is translated to who arranges the service."""
    values: dict[str, str] = {}
    for label, value in kept_items:
        norm_label = _normalize(label)
        for canonical in FIELD_WHITELIST["customs"]:
            norm_canonical = _normalize(canonical)
            if norm_label == norm_canonical or norm_canonical in norm_label or norm_label in norm_canonical:
                values[canonical] = value
                break

    lines = []
    for canonical in FIELD_WHITELIST["customs"]:
        value = values.get(canonical, "N/A")
        if canonical in CUSTOMS_ARRANGE_FIELDS:
            v = value.strip().lower()
            if v.startswith("yes"):
                value = "Jeppesen to arrange"
            elif v.startswith("no"):
                value = "customer will do own"
        lines.append(f"{canonical}: {value}")
    return lines


def format_flight_monitoring_section(kept_items: list[tuple[str, str]]) -> list[str]:
    """Replaces ALL raw Flight Monitoring text with two derived lines:
      Flight Following:  Q2b (send movements as added service) Yes -> "Jeppesen to Provide", No -> "Not Required"
      Flight Monitoring: Q1  (opt-out checkbox) Checked -> "Not Required", Unchecked -> "Flight Following Required"
    Order is fixed (Flight Following first) regardless of extraction order."""
    optout_value = None
    added_service_value = None
    for label, value in kept_items:
        l = label.lower()
        if "opt-out" in l or "opt out" in l or "movement messages" in l:
            optout_value = value.strip().lower()
        elif "added service" in l or "send movements" in l:
            added_service_value = value.strip().lower()

    if added_service_value is not None and added_service_value.startswith("yes"):
        ff_line = "Flight Following: Jeppesen to Provide"
    elif added_service_value is not None and added_service_value.startswith("no"):
        ff_line = "Flight Following: Not Required"
    else:
        ff_line = "Flight Following: (Q2b not answered)"

    if optout_value is not None and optout_value.startswith("check"):
        fm_line = "Flight Monitoring: Not Required"
    elif optout_value is not None and optout_value.startswith("uncheck"):
        fm_line = "Flight Monitoring: Required"
    else:
        fm_line = "Flight Monitoring: (Q1 not answered)"

    return [ff_line, fm_line]


# heading (matched the same way as SECTION_MAP/FIELD_WHITELIST) -> formatter
# that takes the whitelisted (label, value) items for that heading and
# returns the final list of output lines, bypassing the generic
# "[Heading]\nlabel: value" rendering entirely.
SPECIAL_SECTION_FORMATTERS = {
    "flight monitoring": format_flight_monitoring_section,
    "aircraft details": format_aircraft_details_section,
    "customs": format_customs_section,
    "permits": format_permits_section,
    "catering": format_catering_section,
    "ground handling": format_ground_handling_section,
    "company details": format_company_details_section,
    "hotels": format_hotels_section,
    "ground transport": format_ground_transport_section,
    "fuel": format_fuel_section,
}

# Master on/off switch for which subjects are allowed to show extracted
# content in the output. SECTION_MAP/FIELD_WHITELIST stay intact for every
# subject so nothing has to be re-built later -- to bring a subject back,
# just add it here. All other subjects are forced blank regardless of
# what was extracted for them (High Profile Notes' and General Info's
# static boilerplate lines still get added, since those aren't gated by
# this -- see add_static_content()).
ACTIVE_CONTENT_SUBJECTS = {
    "High Profile Notes",
    "General Info",
    "Customs Information",
    "Flight Following",
    "Permits",
    "Catering",
    "Ground Handling",
    "Company Details",
    "Communications",
    "Hotels",
    "Ground Transport",
    "Fuel",
}


def map_to_subjects(sections: list[dict], salesperson: str | None = None) -> dict[str, list[str]]:
    """{subject: [formatted comment lines]}, grouped and labeled by
    source heading so provenance stays traceable. `salesperson` is the
    name selected in the intake UI, threaded through to General Info's
    "Jeppesen FF Account Exec" line -- see add_static_content()."""
    by_subject: dict[str, list[str]] = {s: [] for s in SUBJECTS}
    unmapped_headings = set()
    dropped_fields = []

    for section in sections:
        heading = section.get("heading", "").strip()
        items = section.get("items", [])
        if not items:
            continue
        heading_key = heading.strip().lower()

        # Headings that feed more than one output subject bypass the
        # normal single-subject SECTION_MAP resolution entirely.
        multi_formatters = next(
            (formatters for pattern, formatters in MULTI_SUBJECT_FORMATTERS.items() if pattern in heading_key),
            None,
        )
        if multi_formatters is not None:
            kept_items = []
            for it in items:
                label = it.get("label", "").strip()
                value = it.get("value", "").strip()
                if not is_whitelisted(heading, label):
                    dropped_fields.append(f"{heading} > {label}")
                    continue
                kept_items.append((label, value))
            if kept_items:
                for target_subject, formatter in multi_formatters:
                    lines = formatter(kept_items)
                    if lines:
                        by_subject[target_subject].append("\n".join(lines))
            continue

        subject = resolve_subject(heading)
        if subject is None:
            unmapped_headings.add(heading)
            continue

        kept_items: list[tuple[str, str]] = []
        for it in items:
            label = it.get("label", "").strip()
            value = it.get("value", "").strip()
            if not is_whitelisted(heading, label):
                dropped_fields.append(f"{heading} > {label}")
                continue
            kept_items.append((label, value))
        if not kept_items:
            continue

        special_formatter = next(
            (fn for pattern, fn in SPECIAL_SECTION_FORMATTERS.items() if pattern in heading_key),
            None,
        )
        if special_formatter is not None:
            lines = special_formatter(kept_items)
            block = "\n".join(lines)
        else:
            lines = []
            for label, value in kept_items:
                if any(marker in label.lower() for marker in FLIGHT_TYPE_ITEM_LABELS):
                    lines.append(format_flight_types(value))
                else:
                    lines.append(f"{label}: {value}")
            block = f"[{heading}]\n" + "\n".join(lines)
        by_subject[subject].append(block)

    if unmapped_headings:
        print(
            "  [info] headings with no SECTION_MAP entry (dropped): "
            + ", ".join(sorted(unmapped_headings)),
            file=sys.stderr,
        )
    if dropped_fields:
        print(
            "  [info] fields not on FIELD_WHITELIST (dropped, not extracted): "
            + "; ".join(dropped_fields),
            file=sys.stderr,
        )

    inactive_with_content = [s for s in SUBJECTS if s not in ACTIVE_CONTENT_SUBJECTS and by_subject.get(s)]
    for subject in inactive_with_content:
        by_subject[subject] = []
    if inactive_with_content:
        print(
            "  [info] subjects not in ACTIVE_CONTENT_SUBJECTS (content cleared): "
            + ", ".join(inactive_with_content),
            file=sys.stderr,
        )

    add_static_content(by_subject, salesperson=salesperson)
    return by_subject


def add_static_content(by_subject: dict[str, list[str]], salesperson: str | None = None) -> None:
    """Inject fixed boilerplate that is not derived from the source PDF.
    Runs on every document, appended after any extracted content."""
    by_subject.setdefault("High Profile Notes", []).insert(0, HIGH_PROFILE_NOTES_TEMPLATE)
    by_subject.setdefault("General Info", []).append(
        "\n".join(_general_info_static_lines(salesperson))
    )

    # Fuel's only defined field (Preferred Fuel Provider) always renders
    # with an N/A fallback -- but if it was never answered anywhere in
    # the source doc, no "Fuel" heading chunk has any whitelisted items
    # at all, so the main loop never calls format_fuel_section(). Catch
    # that here rather than showing a blank section.
    if not by_subject.get("Fuel"):
        by_subject.setdefault("Fuel", []).append("\n".join(format_fuel_section([])))


def write_workbook(by_subject: dict[str, list[str]], out_path: Path) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Operational Information"

    header_fill = PatternFill("solid", fgColor="1F3864")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="B7B7B7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    wrap = Alignment(wrap_text=True, vertical="top")

    ws.append(["Subject", "Comments"])
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border

    for subject in SUBJECTS:
        blocks = by_subject.get(subject, [])
        comments = "\n\n".join(blocks)
        ws.append([subject, comments])

    for row in ws.iter_rows(min_row=2):
        row[0].font = Font(bold=True)
        for cell in row:
            cell.border = border
            cell.alignment = wrap

    col_a_width = 26
    col_b_width = 70
    ws.column_dimensions["A"].width = col_a_width
    ws.column_dimensions["B"].width = col_b_width

    # Row height must account for WRAPPED lines, not just explicit "\n"s --
    # a single long line will visually wrap into several lines once
    # word-wrap kicks in at the column width, and undercounting this is
    # exactly what was cutting text off at the bottom of tall cells.
    LINE_HEIGHT_PT = 15.5  # ~single line at default 11pt Calibri
    ROW_PADDING_PT = 6  # a little breathing room top/bottom

    def wrapped_line_count(text: str, width_chars: int) -> int:
        total = 0
        for line in text.split("\n"):
            if not line:
                total += 1
            else:
                total += max(1, -(-len(line) // width_chars))  # ceil division
        return total

    for row in ws.iter_rows(min_row=2):
        subject_lines = wrapped_line_count(row[0].value or "", col_a_width)
        comment_lines = wrapped_line_count(row[1].value or "", col_b_width)
        n_lines = max(subject_lines, comment_lines, 1)
        ws.row_dimensions[row[0].row].height = n_lines * LINE_HEIGHT_PT + ROW_PADDING_PT

    ws.freeze_panes = "A2"

    # Keep each row's Subject + Comments together when printed/exported to
    # PDF. Without this, a wide Comments column gets pushed onto a
    # separate page-group in print view -- i.e. the Subject column prints
    # top-to-bottom first, and the Comments text only shows up later on
    # different pages, looking "blank"/"missing" even though the data is
    # correctly stored in the cell.
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_options.horizontalCentered = True

    wb.save(out_path)


def main():
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("pdf", type=Path, help="Path to the source PDF")
    ap.add_argument("-o", "--out", type=Path, default=None, help="Output .xlsx path")
    ap.add_argument("--use-cache", action="store_true", help="Reuse cached extraction JSON instead of calling the API again")
    args = ap.parse_args()

    out_path = args.out or args.pdf.with_suffix(".operational_info.xlsx")
    cache_path = args.pdf.with_suffix(".extracted.json")

    sections = extract_document(args.pdf, cache_path, use_cache=args.use_cache)
    by_subject = map_to_subjects(sections)
    write_workbook(by_subject, out_path)

    covered = [s for s in SUBJECTS if by_subject.get(s)]
    blank = [s for s in SUBJECTS if not by_subject.get(s)]
    print(f"\nWrote {out_path}")
    print(f"Subjects populated ({len(covered)}): {', '.join(covered)}")
    print(f"Subjects left blank ({len(blank)}): {', '.join(blank)}")


if __name__ == "__main__":
    main()
